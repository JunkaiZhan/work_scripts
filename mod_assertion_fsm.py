#!/Library/Frameworks/Python.framework/Versions/3.7/bin/python3
# Author : zhanjunkai
# Description: Attach the excel form of FSM state transition, and the script
#              will output the assertion and cover for the normal and forbidden conditions

# [normal_list, forbidden_List, cfg_list]
#      |             |
#      |             |------> [ tuple, tuple, tuple ]
#      |
#      |----> [ dist, dist, dist, .... ]
#                |
#                |----> { (source_state, destination_state) -> [condition1, condition2, ...] }

import openpyxl
import itertools
import sys
import os

# read the excel form and get the state transition list
def get_fsm(file_name):

    return_list = []
    normal_list = []
    forbidden_list = []
    transition  = []
    
    wb = openpyxl.load_workbook(file_name)
    ws = wb['Sheet1']
    wss = wb['Sheet2']

    clk_name   = wss.cell(row=2, column=1).value
    rst_name   = wss.cell(row=2, column=2).value
    state_name = wss.cell(row=2, column=3).value
    init_state = wss.cell(row=2, column=4).value
    fsm_name   = wss.cell(row=2, column=5).value
    cfg_list   = [clk_name, rst_name, state_name, init_state, fsm_name]
    
    for row in ws.iter_rows():
        conditions  = []
        fsm_type = row[0].value
        if fsm_type == "NORMAL":
            transition = (row[1].value, row[2].value)
            for _ in row[3:(len(row)+1)]:
                if _.value != None:
                    conditions.append(_.value)
            dict = {transition: conditions}
            normal_list.append(dict)
        elif fsm_type == "FORBIDDEN":
            transition = (row[1].value, row[2].value)
            forbidden_list.append(transition)
    
    print("THE NORMAL STATE TRANSITION LIST IS ========> ")
    for _ in normal_list:
        print(_)

    print("THE FORBIDDEN STATE TRANSITION LIST IS =====> ")
    for _ in forbidden_list:
        print(_)

    print("THE CONFIG VARIABLE NAME IS ================>")
    for _ in cfg_list:
        print(_)

    return_list = [normal_list, forbidden_list, cfg_list]
    return return_list

def write_assertion(fsm_list):

    file_out = open('write_fsm_assertion.sv', 'w')
    
    clk_name = fsm_list[2][0]
    rst_name = fsm_list[2][1]
    state_name = fsm_list[2][2]
    init_state = fsm_list[2][3]
    fsm_name = fsm_list[2][4]

    # write the reset assertion
    out_string  = '    // The property for FSM RESET into the INIT STATE\n'
    out_string += '    property p_reset_for_' + fsm_name + ';\n'
    out_string += '        $rose(' + rst_name + ") |-> [1:3] (" + state_name + ' == ' + init_state + ');\n'
    out_string += '    endproperty\n\n' 

    # write the normal state transition
    normal_list = fsm_list[0]

    for _ in normal_list:

        src = list(_.keys())[0][0]
        dst = list(_.keys())[0][1]
        con = list(_.values())[0]
        
        for i in range(len(con)):
            assert_name = src + "_to_" + dst + "_with_c" + str(i)
            out_string += '    // The property for FSM normal transition from ' + src + ' to ' + dst + ' with condition: ' + con[i] + '\n'
            out_string += '    property p_' + assert_name + ';\n'
            out_string += '        @(posedge ' + clk_name + ');\n'
            out_string += '        disable iff (' + rst_name + ' == 1\'b0)\n'
            out_string += '        ((' + state_name + ' == ' + src + ') && (' + con[i] + ')) |=> (' + state_name + ' == ' + dst + ');\n'
            out_string += '    endproperty\n\n'

        # combine the conditions
        if len(con) > 1:
            # property index
            index = 0
            for i in range(2, len(con)+1):
                for cc in itertools.combinations(con, i):                    
                    # get the assertion name
                    assert_name = src + "_to_" + dst + "_with_comb" + str(index)
                    index += 1
                    cc_list = []                   
                    # get the combination condition string
                    for cc_ in cc:
                        cc_ = "(" + cc_ + ")"
                        cc_list.append(cc_)
                    cc_string = " && ".join(cc_list)                  
                    out_string += '    // The property for FSM normal transition from ' + src + ' to ' + dst + ' with condition: ' + cc_string + '\n'
                    out_string += '    property p_' + assert_name + ';\n'
                    out_string += '        @(posedge ' + clk_name + ');\n'
                    out_string += '        disable iff (' + rst_name + ' == 1\'b0)\n'
                    out_string += '        ((' + state_name + ' == ' + src + ') && (' + cc_string + ')) |=> (' + state_name + ' == ' + dst + ');\n'
                    out_string += '    endproperty\n\n'


    # write the forbidden state transition
    forbidden_list = fsm_list[1]
    for _ in forbidden_list:
        src = _[0]
        dst = _[1]
        assert_name = "forbidden_" + src + "_to_" + dst
        out_string += '    // The property for FSM forbidden transition from ' + src + ' to ' + dst + '\n'
        out_string += '    property p_' + assert_name + ':\n'
        out_string += '        @(posedge ' + clk_name + ');\n'
        out_string += '        disable iff (' + rst_name + ' == 1\'b0)\n'
        out_string += '        not ((' + state_name + ' == ' + src + ') |=> (' + state_name + ' == ' + dst + '));\n'
        out_string += '    endproperty\n\n'

    file_out.write(out_string)
    file_out.close()
       

def main():
    file_name = sys.argv[1]
    if os.path.exists(file_name):
        print("The file name is " + file_name)
        return_list = get_fsm(file_name)
        write_assertion(return_list)
    else:
        print("The file name " + file_name + " is not found !!")
        sys.exit(0)

if __name__ == "__main__":
    main()


